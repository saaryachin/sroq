"""
Sroq Excel Export Module

Excel workbook export from scan results. Mirrors CSV structure exactly.
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


# Column headers (must match CSV export)
_HEADERS = [
    "timestamp",
    "network_name",
    "network_cidr",
    "host_ip",
    "open_ports_count",
    "open_ports",
    "unique_cve_count",
    "max_cvss",
    "critical_count",
    "high_count",
    "medium_count",
    "low_count",
    "unknown_count",
    "risk_score",
    "risk_cvss_sum",
    "top_cves",
]


def _build_rows(results: dict) -> list:
    """
    Build row dicts from scan results (shared with CSV export).

    Returns:
        list of dicts, one per host per network
    """
    rows = []
    timestamp = results.get("timestamp", "")

    for network in results.get("networks", []):
        net_name = network.get("name", "")
        net_cidr = network.get("cidr", "")

        for host in network.get("hosts", []):
            ip = host.get("ip", "")

            open_ports = host.get("open_ports", [])
            open_ports_count = len(open_ports)
            open_ports_str = ";".join(str(p) for p in open_ports)

            vulners = host.get("vulners", {})
            severity = vulners.get("severity", {})
            cves = vulners.get("cves", [])

            top_cves_str = ";".join(
                f"{c['id']}({c['cvss']})" for c in cves[:5]
            )

            rows.append({
                "timestamp":       timestamp,
                "network_name":    net_name,
                "network_cidr":    net_cidr,
                "host_ip":         ip,
                "open_ports_count": open_ports_count,
                "open_ports":      open_ports_str,
                "unique_cve_count": vulners.get("unique_cve_count", 0),
                "max_cvss":        vulners.get("max_cvss", 0.0),
                "critical_count":  severity.get("critical", 0),
                "high_count":      severity.get("high", 0),
                "medium_count":    severity.get("medium", 0),
                "low_count":       severity.get("low", 0),
                "unknown_count":   severity.get("unknown", 0),
                "risk_score":      host.get("risk_score", 0),
                "risk_cvss_sum":   host.get("risk_cvss_sum", 0.0),
                "top_cves":        top_cves_str,
            })

    return rows


def generate_excel(results: dict, output_dir: str) -> str:
    """
    Write an Excel workbook from scan results.

    Mirrors the CSV structure exactly: one row per host per network,
    same columns. Includes frozen header row and auto-filter.

    Args:
        results: Scan results dict
        output_dir: Directory to write the Excel file into

    Returns:
        str: Path to the written .xlsx file, or "" if there are no hosts
    """
    rows = _build_rows(results)

    if not rows:
        return ""

    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    timestamp = results.get("timestamp", "")
    filepath = out_path / f"sroq_{timestamp}.xlsx"

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "hosts"

    # Write header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, row_dict in enumerate(rows, start=2):
        for col_idx, header in enumerate(_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header, ""))

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    # Auto-fit column widths (best-effort, capped at 50)
    for col_idx, header in enumerate(_HEADERS, start=1):
        # Estimate width from header length
        width = min(len(header) + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(filepath)

    return str(filepath)
